#!/usr/bin/env python3
"""Certified H-1B labor condition applications, counted per employer.

Called by scripts/build-h1b-sponsors.mjs, which cannot read these files itself: DOL publishes one
~100MB .xlsx per quarter and the backend has no spreadsheet dependency. Python's openpyxl streams
them in read-only mode, so the whole fiscal year costs a few hundred MB of I/O and no memory to
speak of.

Prints one JSON object to stdout: {"<employer name>": {"certified": N, "cities": [...], "states": [...]}}.
Progress goes to stderr.

The geography is not evidence of sponsorship. It exists so scripts/verify-sponsor-matches.mjs can
tell two same-named companies apart: an employer that filed from Hawthorne, California and a job
board posting roles in Hawthorne, California are almost certainly one company, and the Amsterdam
grocer that shares a name with a US filer is not.

WHAT A CERTIFIED LCA IS, and why it counts as evidence: before filing an H-1B petition an employer
must file a Labor Condition Application naming the role, the worksite and the wage, and attest to
paying it. DOL certifying it is that attestation on the record. It is not an approved petition -
that is what the USCIS data holds - so the two are counted separately and the generated file says
which one confirmed each employer.
"""
import json
import re
import sys
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("openpyxl is required: pip3 install openpyxl", file=sys.stderr)
    sys.exit(2)

# H-1B1 is the Chile/Singapore variant of the same programme. Excluded: E-3 (Australia only) and
# H-1B's cap-exempt cousins, which say nothing about whether the employer sponsors an H-1B.
VISA_CLASSES = {"H-1B", "H-1B1 CHILE", "H-1B1 SINGAPORE"}

employers = defaultdict(lambda: {"certified": 0, "cities": set(), "states": set()})
for path in sys.argv[1:]:
    book = openpyxl.load_workbook(path, read_only=True)
    try:
        sheet = book[book.sheetnames[0]]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows)
        try:
            i_status = header.index("CASE_STATUS")
            i_visa = header.index("VISA_CLASS")
            i_employer = header.index("EMPLOYER_NAME")
            i_city = header.index("EMPLOYER_CITY")
            i_state = header.index("EMPLOYER_STATE")
        except ValueError:
            print(f"{path}: unexpected columns, DOL changed the schema", file=sys.stderr)
            sys.exit(1)
        widest = max(i_status, i_visa, i_employer, i_city, i_state)
        scanned = 0
        kept = 0
        for row in rows:
            scanned += 1
            # A ragged trailing row yields a short tuple, and an IndexError two hours into a 400MB
            # parse costs the whole run.
            if len(row) <= widest:
                continue
            name = row[i_employer]
            if not name:
                continue
            if str(row[i_visa]).strip().upper() not in VISA_CLASSES:
                continue
            # DOL's status domain is Certified, Certified - Withdrawn, Denied, Withdrawn. Only the
            # first is counted: a certification later withdrawn is frequently one where the petition
            # was never filed, and six employers here are confirmed on a SINGLE certification, so
            # the difference decides whether they appear on somebody's board at all.
            if str(row[i_status]).strip().upper() != "CERTIFIED":
                continue
            entry = employers[re.sub(r"\s+", " ", str(name)).strip()]
            entry["certified"] += 1
            if row[i_city]:
                entry["cities"].add(str(row[i_city]).strip().upper())
            if row[i_state]:
                entry["states"].add(str(row[i_state]).strip().upper())
            kept += 1
        print(f"{path}: {scanned} rows, {kept} certified H-1B", file=sys.stderr)
    finally:
        book.close()

json.dump(
    {
        name: {
            "certified": data["certified"],
            # Capped: a national employer files from dozens of cities and the verifier only needs
            # enough of them to recognise a match.
            "cities": sorted(data["cities"])[:12],
            "states": sorted(data["states"])[:12],
        }
        for name, data in employers.items()
    },
    sys.stdout,
)
